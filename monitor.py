import sys
import time
from watchdog.observers import Observer
from watchdog.events import *
import openpyxl
import tkinter as tk
import os


class MyHandler(FileSystemEventHandler):
    def __init__(self):
        self.time = 1
        self.file = 2
        self.action = 3
        self.rows = 2
        self.num = 0; 

    def on_modified(self, event):
        action_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        action = "修改"
        modified_log = action_time + ' ' + action + event.src_path
        print(modified_log)
        ws.cell(row=self.rows, column=1).value = action_time
        ws.cell(row=self.rows, column=2).value = action
        ws.cell(row=self.rows, column=3).value = event.src_path
        self.rows += 1
        wb.save(filename)

    def on_created(self, event):
        action_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        action = "创建"
        created_log = action_time + ' ' + action + event.src_path
        print(created_log)
        ws.cell(row=self.rows, column=1).value = action_time
        ws.cell(row=self.rows, column=2).value = action
        ws.cell(row=self.rows, column=3).value = event.src_path
        self.rows += 1
        wb.save(filename)

    def on_moved(self, event):
        action_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        action = "移动"
        moved_log = action_time + ' ' + action + event.src_path
        print(moved_log)
        ws.cell(row=self.rows, column=1).value = action_time
        ws.cell(row=self.rows, column=2).value = action
        ws.cell(row=self.rows, column=3).value = event.src_path
        self.rows += 1
        wb.save(filename)

    def on_deleted(self, event):
        action_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        action = "删除"
        deleted_log = action_time + ' ' + action + event.src_path
        print(deleted_log)
        ws.cell(row=self.rows, column=1).value = action_time
        ws.cell(row=self.rows, column=2).value = action
        ws.cell(row=self.rows, column=3).value = event.src_path
        self.rows += 1
        wb.save(filename)


def mkdir(path):
    folder = os.path.exists(path)

    if not folder:  # 判断是否存在文件夹如果不存在则创建为文件夹
        os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
    else:
        pass


if __name__ == "__main__":

    print(r"""
    [+] Please enter your Destination of monitoring path after the monitor.py
    [+] Example: python monitor.py D:\test
    [+] Default monitoring path is current path if you don't enter any path
    [+] The logs will be saved in D:\Monitoring record
    """)
    
    des_file = sys.argv[1] if len(sys.argv) > 1 else "."
    # 创建目录
    path = r"D:\Monitoring record"
    mkdir(path)

    wb = openpyxl.Workbook()

    # 新建一个excel文件，并且在单元表为"sheet1"的表中写入数据
    ws = wb.create_sheet("sheet1")

    # 调整列宽
    ws.column_dimensions['A'].width = 20.0
    ws.column_dimensions['B'].width = 10.0
    ws.column_dimensions['C'].width = 80.0

    # 在单元格中写入数据
    ws.cell(row=1, column=1).value = "时间"
    ws.cell(row=1, column=2).value = "行为"
    ws.cell(row=1, column=3).value = "文件路径"

    # 日志文件名
    filename = path + "\\\\" + time.strftime("%Y-%m-%d %H-%M-%S", time.localtime()) + ".xlsx"

    event_handler = MyHandler()
    observer = Observer()
    observer.schedule(event_handler, path=des_file, recursive=True)
    observer.start()
    try:
        while True:
            time.sleep(0.1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()
